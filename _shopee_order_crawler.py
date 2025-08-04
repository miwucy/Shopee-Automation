import time
import pandas as pd
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import re
import os
import platform
from cookie_manager import CookieManager, manual_login_helper, auto_login_with_cookies

print("Current working directory:", os.getcwd())

# 配置字典
config_dict = {
    "advanced": {
        "verbose": True,
        "headless": False
    }
}

try:
    from datetime import date
    global file_name
    file_name = str(date.today())[5:]
    workbook = openpyxl.load_workbook(f'~/Downloads/{file_name}.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(title="order", index=0)

# 工具函數
def safe_click_element(element, driver):
    """安全的元素點擊方法，防止被其他元素遮擋"""
    try:
        # 方法1: 先滾動到元素可見位置
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", element)
        time.sleep(0.3)  
        
        # 方法2: 檢查元素是否可見和可點擊
        if element.is_displayed() and element.is_enabled():
            # 方法3: 嘗試直接點擊
            try:
                element.click()
                return True
            except Exception as e:
                print(f"直接點擊失敗: {e}")
        
        # 方法4: 使用JavaScript點擊
        try:
            driver.execute_script("arguments[0].click();", element)
            return True
        except Exception as e:
            print(f"JavaScript點擊失敗: {e}")
        
        # 方法5: 使用ActionChains點擊
        try:
            ActionChains(driver).move_to_element(element).click().perform()
            return True
        except Exception as e:
            print(f"ActionChains點擊失敗: {e}")
        
        # 方法6: 移除可能的遮擋元素
        try:
            # 移除可能的遮擋元素（如彈窗、提示框等）
            driver.execute_script("""
                var overlays = document.querySelectorAll('.modal, .overlay, .popup, .tooltip, .notification, .loading, .spinner');
                overlays.forEach(function(overlay) {
                    if (overlay.style.display !== 'none') {
                        overlay.style.display = 'none';
                    }
                });
            """)
            time.sleep(0.3)
            element.click()
            return True
        except Exception as e:
            print(f"移除遮擋後點擊失敗: {e}")
        
        return False
        
    except Exception as e:
        print(f"安全點擊方法失敗: {e}")
        return False

def wait_for_element_interactable(driver, by, value, timeout=10):
    """等待元素可交互（可見且可點擊）"""
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((by, value))
        )
        return element
    except TimeoutException:
        print(f"元素 {value} 在 {timeout} 秒內不可交互")
        return None

def get_app_root():
    """獲取應用根目錄"""
    return os.path.dirname(os.path.abspath(__file__))

def clean_uc_exe_cache():
    """清理 UC 執行檔快取"""
    try:
        cache_dir = os.path.expanduser("~/.undetected_chromedriver")
        if os.path.exists(cache_dir):
            import shutil
            shutil.rmtree(cache_dir)
            print("已清理 UC 快取")
        return True
    except Exception as e:
        print(f"清理快取失敗: {e}")
        return False

def get_chromedriver_path(webdriver_path):
    """獲取 ChromeDriver 路徑"""
    chromedriver_path = os.path.join(webdriver_path, "chromedriver")
    if platform.system().lower() == "windows":
        chromedriver_path = os.path.join(webdriver_path, "chromedriver.exe")
    return chromedriver_path

def get_uc_options(uc, config_dict, webdriver_path):
    """獲取 UC Chrome 選項"""
    options = uc.ChromeOptions()
    
    # 基本設置
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-blink-features=AutomationControlled')
    
    
    # 設置用戶代理
    options.add_argument('--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    return options

def load_chromdriver_uc(config_dict):
    """加載 undetected_chromedriver"""
    import undetected_chromedriver as uc

    show_debug_message = True       # debug.
    show_debug_message = False      # online

    if config_dict["advanced"]["verbose"]:
        show_debug_message = True

    Root_Dir = get_app_root()
    webdriver_path = os.path.join(Root_Dir, "webdriver")
    chromedriver_path = get_chromedriver_path(webdriver_path)

    if not os.path.exists(webdriver_path):
        os.mkdir(webdriver_path)

    # 嘗試安裝 chromedriver_autoinstaller
    chromedriver_autoinstaller_max = None
    try:
        import chromedriver_autoinstaller
        chromedriver_autoinstaller_max = chromedriver_autoinstaller
    except ImportError:
        print("chromedriver_autoinstaller 未安裝，嘗試安裝...")
        try:
            import subprocess
            subprocess.check_call(["pip", "install", "chromedriver_autoinstaller"])
            import chromedriver_autoinstaller
            chromedriver_autoinstaller_max = chromedriver_autoinstaller
        except Exception as e:
            print(f"安裝 chromedriver_autoinstaller 失敗: {e}")

    if not os.path.exists(chromedriver_path):
        print("ChromeDriver not exist, try to download to:", webdriver_path)
        try:
            if chromedriver_autoinstaller_max is not None:
                chromedriver_autoinstaller_max.install(path=webdriver_path)
            else:
                print("chromedriver_autoinstaller_max not available, please manually download chromedriver")
            if not os.path.exists(chromedriver_path):
                print("check installed chrome version fail, download last known good version.")
                chromedriver_autoinstaller_max.install(path=webdriver_path, detect_installed_version=False)
        except Exception as exc:
            print(exc)
    else:
        print("ChromeDriver exist:", chromedriver_path)

    driver = None
    if os.path.exists(chromedriver_path):
        # use chromedriver_autodownload instead of uc auto download.
        is_cache_exist = clean_uc_exe_cache()

        fail_1 = False
        lanch_uc_with_path = True
        if "macos" in platform.platform().lower():
            if "arm64" in platform.platform().lower():
                lanch_uc_with_path = False

        if lanch_uc_with_path:
            try:
                options = get_uc_options(uc, config_dict, webdriver_path)
                driver = uc.Chrome(driver_executable_path=chromedriver_path, options=options, headless=config_dict["advanced"]["headless"])
            except Exception as exc:
                print(exc)
                error_message = str(exc)
                left_part = None
                if "Stacktrace:" in error_message:
                    left_part = error_message.split("Stacktrace:")[0]
                    print(left_part)

                if "This version of ChromeDriver only supports Chrome version" in error_message:
                    print("Chrome version not match")
                fail_1 = True
        else:
            fail_1 = True

        fail_2 = False
        if fail_1:
            try:
                options = get_uc_options(uc, config_dict, webdriver_path)
                driver = uc.Chrome(options=options)
            except Exception as exc:
                print(exc)
                fail_2 = True

        if fail_2:
            # remove exist chromedriver, download again.
            try:
                print("Deleting exist and download ChromeDriver again.")
                os.unlink(chromedriver_path)
            except Exception as exc2:
                print(exc2)
                pass

            try:
                if chromedriver_autoinstaller_max:
                    chromedriver_autoinstaller_max.install(path=webdriver_path)
                    options = get_uc_options(uc, config_dict, webdriver_path)
                    driver = uc.Chrome(driver_executable_path=chromedriver_path, options=options)
            except Exception as exc2:
                print(exc2)
                pass
    else:
        print("WebDriver not found at path:", chromedriver_path)

    if driver is None:
        print('WebDriver object is still None..., try download by uc.')
        try:
            options = get_uc_options(uc, config_dict, webdriver_path)
            driver = uc.Chrome(options=options)
        except Exception as exc:
            print(exc)
            error_message = str(exc)
            left_part = None
            if "Stacktrace:" in error_message:
                left_part = error_message.split("Stacktrace:")[0]
                print(left_part)

            if "This version of ChromeDriver only supports Chrome version" in error_message:
                print("Chrome version not match")
            pass

    if driver is None:
        print("create web drive object by undetected_chromedriver fail!")
        print("建議您自行下載 ChromeDriver 到 webdriver 的資料夾下")
        print("you need manually download ChromeDriver to webdriver folder.")

    return driver

def transition_time(original_date):
    from datetime import datetime
    formatted_date = datetime.strptime(original_date, "%y%m%d").strftime("%Y/%m/%d")
    return formatted_date


def transition_delivery_way(text):
    if text == '蝦皮店到店':
        return '蝦皮'
    elif '7' in text:
        return '7-11'
    elif 'ok' in text.lower():
        return 'OK'
    return text

def filter_emoji(desstr, restr=''):
    try:
        co = re.compile(u'[\U00010000-\U0010ffff]')
    except re.error:
        co = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
    return co.sub(restr, desstr)

def write_order_xlsx(datas):
    # 將數據寫入一行
    for data in datas:
        sheet.append(data)

    current_row = sheet.max_row
    # print(current_row)
    for col in range(1, 7):
        sheet.merge_cells(start_row=max(current_row - len(datas) + 1, 1), start_column=col, end_row=current_row,
                          end_column=col)
    for col in range(15, 16):
        sheet.merge_cells(start_row=max(current_row - len(datas) + 1, 1), start_column=col, end_row=current_row,
                          end_column=col)
    sheet.merge_cells(start_row=max(1, current_row - len(datas) + 1), start_column=10, end_row=current_row,
                      end_column=10)
    for row in sheet.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=15):
        for cell in row:
            cell.font = Font(name=u'微軟正黑', size=10)
            cell.alignment = Alignment(vertical='center', horizontal='center')
    downloads_path = os.path.expanduser('~/Downloads')
    file_path = os.path.join(downloads_path, file_name + '.xlsx')
    workbook.save(file_path) 





def process_order(order, driver, wait):
    """處理單個訂單"""
    try:
        # 獲取訂單基本信息
        order_card_header = WebDriverWait(order, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'order-card-header')),
            "order-card-header Not Found"
        )
        
        # 提取訂單日期
        date = WebDriverWait(order, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "order-sn")),
            "order-sn Not Found"
        ).text[5:11]
        date = transition_time(date)
        
        # 提取買家ID
        buyer_id = WebDriverWait(order, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'buyer-username')),
            "buyer-username Not Found"
        ).text
        
        # 提取配送方式
        delivery_way = WebDriverWait(order, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "fulfilment-channel-name")),
            "fulfilment-channel-name Not Found"
        ).text
        delivery_way = transition_delivery_way(delivery_way)
        
        # 獲取訂單卡片主體元素並點擊
        total_price_page = WebDriverWait(order, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "order-card-body")),
            "total-price Not Found"
        )
        
        # 使用安全點擊方法
        if not safe_click_element(total_price_page, driver):
            print("❌ 無法點擊訂單卡片，跳過此訂單")
            return False
        
        # 切換到新窗口獲取總價
        driver.switch_to.window(driver.window_handles[-1])
        total_price = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'amount')),
            "income-price Not Found"
        ).text 
        total_price = int(total_price[3:].replace(',', ''))
        driver.close()
        driver.switch_to.window(driver.window_handles[-1])
        
        # 獲取商品信息
        items = order.find_elements(By.CLASS_NAME, 'item-info')
        data = []
        
        for item in items:
            # 提取商品名稱
            item_name = WebDriverWait(item, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'item-name')),
                "item-name elements not found"
            ).text
            item_name = filter_emoji(item_name)
            
            # 提取商品描述（顏色、尺寸）
            color, size = "", ""
            try:
                print("尋找item-description...")
                item_description = WebDriverWait(item, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'item-description')),
                    "item-description elements not found"
                ).text[6:]
                if ',' in item_description:
                    color, size = item_description.split(',')
                elif item_description.isnumeric():
                    size = item_description
                else:
                    color = item_description
            except TimeoutException:
                item_description = ""
            
            # 提取商品數量
            item_amount = WebDriverWait(item, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'item-amount')),
                "item-amount elements not found"
            ).text
            item_amount = int(item_amount[1:])  # remove "x"
            
            # 根據數量創建數據條目
            note = ""
            for _ in range(item_amount):
                data.append([
                    date, buyer_id, "", "", delivery_way, "", item_name, color, size,
                    total_price, None, None, None, None, note
                ])
        
        # 寫入Excel
        write_order_xlsx(data)
        return True
        
    except NoSuchElementException as e:
        print(f'Element not found within the given time: {e}')
        return False
    except Exception as e:
        print(f'處理訂單時發生錯誤: {e}')
        return False

def wait_for_orders_to_load(wrapper, wait):
    """等待訂單列表加載完成"""
    orders = wrapper.find_elements(By.CLASS_NAME, 'order-card')
    
    while len(orders) == 0: 
        print(f"⏳ 訂單數量為0，等待頁面完全加載...")
        time.sleep(2)  # 等待2秒
        orders = wrapper.find_elements(By.CLASS_NAME, 'order-card')  
        print(f"🔄 重新檢查：找到 {len(orders)} 筆訂單")
    
    print(f"✅ 確認找到 {len(orders)} 筆訂單，開始處理...")
    return orders

def handle_login(driver, cookie_manager, wait):
    """處理登錄流程"""
    if not auto_login_with_cookies(driver, cookie_manager):
        # 如果 cookies 登錄失敗，進行手動登錄
        driver.get('https://seller.shopee.tw/portal/sale/shipment?type=toship&source=all&sort_by=create_date_desc')
        print("等待頁面加載...")
        time.sleep(5)
        
        # 檢查是否需要登錄
        try:
            login_button = wait.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '登入')]")))
            print("檢測到登入按鈕，需要手動登錄")
            manual_login_helper(driver, cookie_manager)
        except TimeoutException:
            print("未檢測到登入按鈕，可能已經登錄")
    else:
        print("✅ 使用cookies自動登錄成功")

def wait_for_page_ready(driver, wait):
    """等待頁面準備就緒"""
    while True:
        try:
            # 嘗試找到元素
            order_list_table_shipment = wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, 'order-list-table-shipment')))
            print("✅ 頁面元素已找到，可以進行後續操作")
            break  # 如果成功找到元素，跳出迴圈
        except TimeoutException:
            print("⚠️ 請先登入")
            time.sleep(5)  # 休眠5秒再次檢查，避免過於頻繁的請求

def main():
    """主函數"""
    
    print("🚀 開始執行蝦皮訂單爬蟲...")
    
    # 1. 初始化 WebDriver
    print("正在初始化 undetected_chromedriver...")
    driver = load_chromdriver_uc(config_dict)
    
    if driver is None:
        print("❌ 無法創建 WebDriver，程序退出")
        return
    
    print("✅ WebDriver 初始化成功")
    
    try:
        # 2. 初始化 cookie 管理器和等待器
        cookie_manager = CookieManager()
        wait = WebDriverWait(driver, 10)
        
        # 3. 處理登錄
        handle_login(driver, cookie_manager, wait)
        
        # 4. 等待頁面準備就緒
        wait_for_page_ready(driver, wait)
        
        # 5. 等待訂單列表容器加載完成
        wrapper = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'table-body-wrapper')))
        print("✅ 訂單列表容器已加載")
        
        # 6. 等待訂單數據加載
        orders = wait_for_orders_to_load(wrapper, wait)
        
        processed_count = 0
        
        for i, order in enumerate(reversed(orders), 1):
            print(f"🔄 處理第 {i}/{len(orders)} 筆訂單...")
            if process_order(order, driver, wait):
                processed_count += 1
                print(f"✅ 第 {i} 筆訂單處理完成")
            else:
                print(f"❌ 第 {i} 筆訂單處理失敗")
        
        print(f"🎉 爬蟲完成！成功處理 {processed_count}/{len(orders)} 筆訂單")
        
    except Exception as e:
        print(f"❌ 程序執行過程中發生錯誤: {e}")
    finally:
        # 8. 清理資源
        if driver:
            driver.quit()
            print("✅ WebDriver 已關閉")

if __name__ == "__main__":
    main()